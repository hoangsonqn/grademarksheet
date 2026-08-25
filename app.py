import io
import json, tempfile, os
from pathlib import Path
import streamlit as st
import pandas as pd
import numpy as np
import cv2
import fitz
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment, PatternFill

BASE_DIR = Path(__file__).resolve().parent
ANSWER_KEYS_PATH = BASE_DIR / "answers.json"
TEMPLATE_PATH = BASE_DIR / "template.json"
MARKS_AVAILABLE = [5, 10, 15, 20, 25, 30, 35, 40, 45, 50]
QUESTIONS_PER_MARK = 80
VALID_ANSWERS = {"A", "B", "C", "D"}

@st.cache_data
def load_default_template():
    with TEMPLATE_PATH.open("r", encoding="utf-8") as f:
        return json.load(f)

@st.cache_data
def load_answer_config():
    """Load answer keys and optional per-mark instructions once."""
    if not ANSWER_KEYS_PATH.exists():
        return {}, {}

    with ANSWER_KEYS_PATH.open("r", encoding="utf-8") as f:
        data = json.load(f)

    if not isinstance(data, dict):
        raise ValueError("Nội dung answers.json phải là một object JSON.")

    raw_instructions = data.get("_instructions", {})
    if not isinstance(raw_instructions, dict):
        raise ValueError("Mục _instructions trong answers.json phải là một object JSON.")

    instructions = {}
    for mark_text, raw_instruction in raw_instructions.items():
        try:
            mark = int(mark_text)
        except (TypeError, ValueError) as exc:
            raise ValueError(f"Mốc hướng dẫn không hợp lệ: {mark_text!r}.") from exc

        instruction = str(raw_instruction).strip()
        if instruction:
            instructions[mark] = instruction

    all_keys = {}
    for mark_text, raw_key in data.items():
        # Các khóa bắt đầu bằng '_' là metadata, không phải mốc đề.
        if str(mark_text).startswith("_"):
            continue

        try:
            mark = int(mark_text)
        except (TypeError, ValueError) as exc:
            raise ValueError(f"Mốc câu hỏi không hợp lệ: {mark_text!r}.") from exc

        if not isinstance(raw_key, dict):
            raise ValueError(f"Đáp án mốc {mark} phải là một object JSON.")

        answer_key = {}
        for question_text, raw_answer in raw_key.items():
            try:
                question = int(question_text)
            except (TypeError, ValueError) as exc:
                raise ValueError(
                    f"Số câu {question_text!r} tại mốc {mark} không hợp lệ."
                ) from exc

            answer = str(raw_answer).strip().upper()
            if answer not in VALID_ANSWERS:
                raise ValueError(
                    f"Đáp án câu {question} tại mốc {mark} phải là A, B, C hoặc D."
                )
            answer_key[question] = answer

        expected_questions = set(range(1, QUESTIONS_PER_MARK + 1))
        if set(answer_key) != expected_questions:
            raise ValueError(
                f"Mốc {mark} phải có đủ và đúng {QUESTIONS_PER_MARK} câu "
                f"từ 1 đến {QUESTIONS_PER_MARK}."
            )
        all_keys[mark] = answer_key

    return all_keys, instructions

def normalize_student_number(value):
    """Normalize roster/OMR numbers so values such as 01 and 1 match."""
    if value is None or pd.isna(value):
        return None

    text = str(value).strip()
    if text.isdigit():
        return str(int(text))
    if text.endswith(".0") and text[:-2].isdigit():
        return str(int(text[:-2]))
    return None

@st.cache_data
def load_student_roster(file_name, file_bytes):
    """Read a two-column roster and return a normalized STT-to-name mapping."""
    suffix = Path(file_name).suffix.lower()
    source = io.BytesIO(file_bytes)

    if suffix == ".xlsx":
        df = pd.read_excel(source, header=None, dtype=str)
    elif suffix == ".csv":
        df = None
        for encoding in ("utf-8-sig", "utf-8", "cp1258"):
            try:
                source.seek(0)
                df = pd.read_csv(
                    source,
                    header=None,
                    dtype=str,
                    encoding=encoding,
                    sep=None,
                    engine="python",
                    keep_default_na=False,
                )
                break
            except UnicodeDecodeError:
                continue
        if df is None:
            raise ValueError("File CSV phải dùng bảng mã UTF-8 hoặc Windows-1258.")
    else:
        raise ValueError("Danh sách lớp phải là file .xlsx hoặc .csv.")

    if df.shape[1] < 2:
        raise ValueError("Danh sách lớp phải có 2 cột: Số thứ tự và Họ tên.")

    roster = {}
    for row_index, row in df.iloc[:, :2].iterrows():
        raw_number = "" if pd.isna(row.iloc[0]) else str(row.iloc[0]).strip()
        student_name = "" if pd.isna(row.iloc[1]) else str(row.iloc[1]).strip()

        if not raw_number and not student_name:
            continue

        student_number = normalize_student_number(raw_number)
        if student_number is None:
            header_text = f"{raw_number} {student_name}".lower()
            header_tokens = (
                "stt", "số tt", "số thứ tự", "họ tên", "họ và tên", "name"
            )
            if not roster and any(token in header_text for token in header_tokens):
                continue
            raise ValueError(f"STT không hợp lệ tại dòng {row_index + 1}: {raw_number!r}.")

        if not student_name:
            raise ValueError(f"Thiếu họ tên tại dòng {row_index + 1}.")
        if student_number in roster:
            raise ValueError(f"STT {raw_number} bị trùng trong danh sách lớp.")
        roster[student_number] = student_name

    if not roster:
        raise ValueError("Danh sách lớp không có học viên hợp lệ.")
    return roster

DARK_THRESHOLD = 215
MIN_GAP_TO_2ND = 18

def find_squares(gray):
    H, W = gray.shape[:2]
    area_total = H * W
    _, th = cv2.threshold(gray, 100, 255, cv2.THRESH_BINARY_INV)
    contours, _ = cv2.findContours(th, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    squares = []
    for c in contours:
        x, y, w, h = cv2.boundingRect(c)
        area = w * h
        if (area_total * 0.00012) < area < (area_total * 0.0020) and 0.7 < w / h < 1.35:
            squares.append((x + w / 2.0, y + h / 2.0, w, h))
    return squares

def detect_corners(gray):
    H, W = gray.shape[:2]
    squares = find_squares(gray)
    corners = {'TL': None, 'TR': None, 'BL': None, 'BR': None}
    best = {'TL': 1e18, 'TR': 1e18, 'BL': 1e18, 'BR': 1e18}
    for (cx, cy, w, h) in squares:
        d = {'TL': cx**2 + cy**2, 'TR': (W - cx)**2 + cy**2, 'BL': cx**2 + (H - cy)**2, 'BR': (W - cx)**2 + (H - cy)**2}
        for key, val in d.items():
            if val < best[key]:
                best[key] = val
                corners[key] = (cx, cy)
    return corners

def warp_to_template(page_img, template):
    gray = cv2.cvtColor(page_img, cv2.COLOR_BGR2GRAY)
    corners = detect_corners(gray)
    if any(v is None for v in corners.values()):
        return None, "Không tìm thấy đủ 4 điểm neo ở góc phiếu."
    src = np.array([corners['TL'], corners['TR'], corners['BL'], corners['BR']], dtype=np.float32)
    tc = template['corners']
    dst = np.array([tc['TL'], tc['TR'], tc['BL'], tc['BR']], dtype=np.float32)
    M = cv2.getPerspectiveTransform(src, dst)
    W, H = template['page_size']
    return cv2.warpPerspective(page_img, M, (int(W), int(H)), flags=cv2.INTER_LINEAR, borderValue=(255, 255, 255)), None

def circle_darkness(gray, x, y, r):
    x, y, r = int(round(x)), int(round(y)), int(r)
    H, W = gray.shape[:2]
    patch = gray[max(0, y-r):min(H, y+r+1), max(0, x-r):min(W, x+r+1)]
    yy, xx = np.ogrid[max(0, y-r):min(H, y+r+1), max(0, x-r):min(W, x+r+1)]
    mask = (xx - x)**2 + (yy - y)**2 <= r*r
    return float(patch[mask].mean()) if mask.sum() > 0 else 255.0

def read_answer(gray, options_dict, radius):
    vals = {opt: circle_darkness(gray, xy[0], xy[1], radius) for opt, xy in options_dict.items()}
    ordered = sorted(vals.items(), key=lambda kv: kv[1])
    darkest_opt, darkest_val = ordered[0]
    second_val = ordered[1][1] if len(ordered) > 1 else 255
    if darkest_val > DARK_THRESHOLD: return None, False
    if (second_val - darkest_val) < MIN_GAP_TO_2ND and second_val <= DARK_THRESHOLD: return None, True
    return darkest_opt, False

def read_digit_column(gray, rows, col_x, radius):
    vals = [circle_darkness(gray, col_x, y, radius) for y in rows]
    idx = int(np.argmin(vals))
    if vals[idx] > DARK_THRESHOLD:
        return None
    ordered = sorted(vals)
    if len(ordered) > 1 and (ordered[1] - ordered[0]) < MIN_GAP_TO_2ND and ordered[1] <= DARK_THRESHOLD:
        return None
    return idx

def read_stt(gray, stt_template, radius):
    if not stt_template:
        return None
    cols = stt_template.get("cols", [])
    if not cols and "col" in stt_template:
        cols = [stt_template["col"]]
    rows = stt_template.get("rows", [])
    digits = []
    for col_x in cols:
        d = read_digit_column(gray, rows, col_x, radius)
        if d is None:
            return None
        digits.append(str(d))
    return "".join(digits)

def read_made(gray, made_template, radius):
    if not made_template: return None
    digits = [read_digit_column(gray, r, c, radius) for c, r in zip(made_template['cols'], made_template['rows_per_col'])]
    return "".join(str(d) for d in digits) if None not in digits else None

def grade_page(page_img, template, answer_key, n_questions):
    warped, err = warp_to_template(page_img, template)
    result = {"ok": err is None, "error": err, "stt": None, "made": None,
              "dung": 0, "sai": 0, "bo_trong": 0, "khong_hop_le": 0, "diem": None, "name_crop": None}
    if err: return result

    gray = cv2.cvtColor(warped, cv2.COLOR_BGR2GRAY)
    radius = 12

    result["stt"] = read_stt(gray, template.get("stt_template"), radius)
    result["made"] = read_made(gray, template.get("made_template"), radius)

    dung = sai = bo_trong = khong_hop_le = 0
    bubble_radius = template.get("bubble_radius", 15) - 3
    for q in range(1, n_questions + 1):
        opts = template["questions"].get(str(q))
        if not opts: continue
        ans, is_multi = read_answer(gray, opts, bubble_radius)
        correct = answer_key.get(q)
        if is_multi: khong_hop_le += 1
        elif ans is None: bo_trong += 1
        elif correct is not None and ans == correct: dung += 1
        else: sai += 1

    result.update({"dung": dung, "sai": sai, "bo_trong": bo_trong, "khong_hop_le": khong_hop_le,
                   "diem": round(dung / n_questions * 100, 2) if n_questions else None})
    x0r, y0r, x1r, y1r = template["name_crop_rel"]
    W, H = template["page_size"]
    result["name_crop"] = warped[int(y0r*H):int(y1r*H), int(x0r*W):int(x1r*W)]
    return result

def export_excel(results, out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Ket qua"
    headers = ["Trang", "Ảnh (Tên/Lớp)", "STT", "Họ và tên", "Mã đề", "Số câu đúng", "Số câu sai", "Bỏ trống/Lỗi", "Điểm (/100)", "Ghi chú"]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DDEBF7")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 32
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 15
    ws.column_dimensions["I"].width = 12
    ws.column_dimensions["J"].width = 38

    tmp_files = []
    for i, r in enumerate(results, start=2):
        ws.row_dimensions[i].height = 70
        ws.cell(row=i, column=1, value=r["page"])
        ws.cell(row=i, column=3, value=r["stt"] if r["stt"] is not None else "")
        ws.cell(row=i, column=4, value=r.get("student_name", ""))
        ws.cell(row=i, column=5, value=r["made"] if r["made"] else "")
        ws.cell(row=i, column=6, value=r["dung"])
        ws.cell(row=i, column=7, value=r["sai"])
        ws.cell(row=i, column=8, value=r["bo_trong"] + r["khong_hop_le"])
        ws.cell(row=i, column=9, value=r["diem"])
        notes = [r.get("error"), r.get("roster_note")]
        ws.cell(row=i, column=10, value="; ".join(note for note in notes if note))

        if r["name_crop"] is not None:
            fd, tmp_path = tempfile.mkstemp(suffix=".png")
            os.close(fd)
            cv2.imwrite(tmp_path, r["name_crop"])
            tmp_files.append(tmp_path)
            img = XLImage(tmp_path)
            img.width = int(img.width * (90 / img.height))
            img.height = 90
            ws.add_image(img, f"B{i}")

    wb.save(out_path)
    for p in tmp_files:
        try: os.remove(p)
        except OSError: pass

# --- Streamlit UI ---
st.set_page_config(page_title="Chấm Thi Trắc Nghiệm Tự Động", layout="wide")
st.title("📝 Hệ Thống Chấm Điểm Trắc Nghiệm Tự Động")

try:
    all_keys, all_instructions = load_answer_config()
    answer_keys_error = None
except (OSError, json.JSONDecodeError, ValueError) as exc:
    all_keys = {}
    all_instructions = {}
    answer_keys_error = str(exc)

with st.sidebar:
    st.header("⚙️ Cấu hình chấm bài")
    selected_mark = st.selectbox(
        "🎯 Chọn mốc đề",
        MARKS_AVAILABLE,
        index=0,
    )
    answer_key = all_keys.get(selected_mark, {})
    selected_instruction = all_instructions.get(selected_mark)

    if answer_keys_error:
        st.error(f"Không thể đọc answers.json: {answer_keys_error}")
    elif answer_key:
        st.success(
            f"Đã nạp đáp án mốc **{selected_mark}**: "
            f"**{len(answer_key)} câu**"
        )
    else:
        st.warning(f"Chưa có dữ liệu đáp án cho mốc {selected_mark}.")

    if selected_instruction:
        st.info(f"📌 **Hướng dẫn:** {selected_instruction}")

    st.divider()
    roster_file = st.file_uploader(
        "👥 Tải danh sách lớp",
        type=["xlsx", "csv"],
        help="File gồm 2 cột: Số thứ tự và Họ tên. Có thể có hàng tiêu đề.",
    )
    student_roster = {}
    roster_error = None
    if roster_file is not None:
        try:
            student_roster = load_student_roster(
                roster_file.name,
                roster_file.getvalue(),
            )
        except (ValueError, OSError) as exc:
            roster_error = str(exc)
            st.error(f"Không thể đọc danh sách lớp: {roster_error}")
        else:
            st.success(f"Đã nạp **{len(student_roster)} học viên**")

pdf_file = st.file_uploader("📥 Tải lên file PDF bài làm", type=["pdf"])

if st.button("🚀 Bắt đầu chấm điểm", type="primary"):
    if not pdf_file:
        st.error("Vui lòng tải lên file PDF bài làm.")
    elif roster_file is None:
        st.error("Vui lòng tải lên danh sách lớp.")
    elif roster_error:
        st.error("Danh sách lớp chưa hợp lệ. Vui lòng sửa file và tải lại.")
    elif not answer_key:
        st.error(
            f"Không tìm thấy đáp án cho mốc {selected_mark}. "
            "Vui lòng kiểm tra lại answers.json."
        )
    else:
        template = load_default_template()
        n_questions = len(answer_key)

        missing_template_questions = [
            q for q in range(1, n_questions + 1)
            if str(q) not in template.get("questions", {})
        ]
        if missing_template_questions:
            st.error(
                "Template phiếu thiếu tọa độ cho các câu: "
                + ", ".join(map(str, missing_template_questions))
            )
            st.stop()

        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp_pdf:
            tmp_pdf.write(pdf_file.read())
            tmp_pdf_path = tmp_pdf.name

        doc = fitz.open(tmp_pdf_path)
        total_pages = len(doc)
        progress_bar = st.progress(0)
        status_text = st.empty()

        results = []
        for i, page in enumerate(doc):
            pw = page.rect.width
            zoom = template["page_size"][0] / pw if pw else 1.0
            pix = page.get_pixmap(matrix=fitz.Matrix(zoom, zoom), alpha=False)
            arr = np.frombuffer(pix.samples, dtype=np.uint8).reshape(pix.height, pix.width, pix.n)
            img = cv2.cvtColor(arr, cv2.COLOR_RGBA2BGR if pix.n == 4 else cv2.COLOR_RGB2BGR)

            res = grade_page(img, template, answer_key, n_questions)
            res["page"] = i + 1
            normalized_stt = normalize_student_number(res["stt"])
            res["student_name"] = (
                student_roster.get(normalized_stt, "")
                if normalized_stt is not None
                else ""
            )
            res["roster_note"] = ""
            if normalized_stt is not None and not res["student_name"]:
                res["roster_note"] = (
                    f"Không tìm thấy STT {res['stt']} trong danh sách lớp"
                )
            results.append(res)
            
            progress_bar.progress((i + 1) / total_pages)
            status_text.text(f"Đang chấm trang {i + 1}/{total_pages}...")

        doc.close()
        os.remove(tmp_pdf_path)
        out_excel = "ket_qua_cham.xlsx"
        export_excel(results, out_excel)
        status_text.success("✅ Đã hoàn tất chấm điểm!")

        summary_data = []
        for r in results:
            summary_data.append({
                "Trang": r["page"],
                "STT": r["stt"] or "Không nhận diện được",
                "Họ và tên": r.get("student_name", "") or "Không tìm thấy",
                "Mã đề": r["made"] or "-",
                "Số câu đúng": r["dung"],
                "Số câu sai": r["sai"],
                "Lỗi/Trống": r["bo_trong"] + r["khong_hop_le"],
                "Điểm": r["diem"],
                "Trạng thái": (
                    r["error"] or r.get("roster_note") or "Thành công"
                )
            })
        st.dataframe(pd.DataFrame(summary_data), use_container_width=True)

        unmatched_count = sum(1 for r in results if r.get("roster_note"))
        if unmatched_count:
            st.warning(
                f"Có {unmatched_count} bài không tìm thấy STT tương ứng "
                "trong danh sách lớp."
            )

        with open(out_excel, "rb") as f:
            st.download_button("📥 Tải về file Excel kết quả chi tiết", data=f, file_name="Ket_qua_cham_trac_nghiem.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
